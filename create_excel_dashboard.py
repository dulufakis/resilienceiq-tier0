#!/usr/bin/env python3
"""
ResilienceIQ Tier 0 — Excel Dashboard Generator (Real Data)
Fetches live data from Eurostat, Google Trends, Open-Meteo and builds
a branded .xlsx with KPIs, charts, and conditional formatting.
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd
import numpy as np
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule

from data_fetcher import (
    WEIGHTS_FALLBACK, COMPONENT_NAMES, fetch_eurostat_tourism, fetch_google_trends,
    fetch_weather, build_resilience_snapshot,
)

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)


# ── Styles ──
BLUE = "1E50B4"
LIGHT_BLUE = "E8F0FE"
GREEN = "27AE60"
ORANGE = "E67E22"
RED = "C0392B"
WHITE = "FFFFFF"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=BLUE, size=16)
subtitle_font = Font(name="Calibri", bold=True, color=BLUE, size=12)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
center = Alignment(horizontal="center", vertical="center")


def build_excel(df: pd.DataFrame, df_euro: pd.DataFrame, output_path: str):
    wb = Workbook()

    # ============================================================
    # SHEET 1: Dashboard
    # ============================================================
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = BLUE

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "ResilienceIQ - Tourism Resilience Dashboard (Real Data)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Sources: Eurostat, Google Trends, Open-Meteo"
    ws["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Municipality", "Region", "Arrivals", "Nights", "Avg Stay",
               "Demand Idx", "Weather", "Score", "Zone"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), 5):
        vals = [
            row["dimos"], row["region"],
            row.get("arrivals"), row.get("nights"),
            row.get("avg_stay"), row.get("demand_index"),
            row.get("weather_score"), row["resilience_score"], row["zone"],
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(v):
                cell.value = "N/A"
            elif isinstance(v, float) and col_idx in (3, 4):
                cell.value = round(v)
                cell.number_format = "#,##0"
            elif isinstance(v, float):
                cell.value = round(v, 1)
            else:
                cell.value = v
            cell.alignment = center
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    last_row = 4 + len(df)

    # Column widths
    for i, w in enumerate([16, 16, 14, 16, 10, 12, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Zone thresholds (μ ± 0.5σ)
    mu = df["resilience_score"].mean()
    sigma = df["resilience_score"].std()
    high_thresh = mu + 0.5 * sigma
    low_thresh = mu - 0.5 * sigma

    # Conditional formatting on Score (col H)
    score_range = f"H5:H{last_row}"
    ws.conditional_formatting.add(score_range,
        CellIsRule(operator="greaterThanOrEqual", formula=[str(round(high_thresh, 2))], fill=PatternFill(bgColor=GREEN)))
    ws.conditional_formatting.add(score_range,
        CellIsRule(operator="between", formula=[str(round(low_thresh, 2)), str(round(high_thresh - 0.01, 2))], fill=PatternFill(bgColor=ORANGE)))
    ws.conditional_formatting.add(score_range,
        CellIsRule(operator="lessThan", formula=[str(round(low_thresh, 2))], fill=PatternFill(bgColor=RED)))
    ws.conditional_formatting.add(score_range,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color=BLUE))

    # Bar Chart
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Resilience Score by Municipality"
    chart1.y_axis.title = "Score"
    chart1.style = 10

    data_ref = Reference(ws, min_col=8, min_row=4, max_row=last_row)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=last_row)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.width = 22
    chart1.height = 14

    series = chart1.series[0]
    for i, (_, row) in enumerate(df.iterrows()):
        pt = DataPoint(idx=i)
        color = GREEN if row["resilience_score"] >= high_thresh else ORANGE if row["resilience_score"] >= low_thresh else RED
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)
    series.graphicalProperties.line.noFill = True
    chart1.legend = None
    ws.add_chart(chart1, f"A{last_row + 2}")

    # ============================================================
    # SHEET 2: Component Analysis
    # ============================================================
    ws2 = wb.create_sheet("Component Analysis")
    ws2.sheet_properties.tabColor = GREEN

    ws2.merge_cells("A1:F1")
    ws2["A1"] = "Normalized Component Scores (0-100)"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center")

    comp_headers = ["Municipality", "Arrivals", "Nights", "Demand", "Weather", "Avg Stay"]
    for col_idx, h in enumerate(comp_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        vals = [row["dimos"], row["norm_arrivals"], row["norm_nights"],
                row["norm_demand"], row["norm_weather"], row["norm_avg_stay"]]
        for col_idx, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.value = round(v, 1) if isinstance(v, (float, np.floating)) else v
            cell.alignment = center
            cell.border = thin_border

    comp_last = 3 + len(df)
    for i, w in enumerate([16, 12, 12, 12, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Radar Chart
    radar = RadarChart()
    radar.type = "filled"
    radar.title = "Resilience Profile (Top 5)"
    radar.style = 26
    radar.width = 22
    radar.height = 16

    cats = Reference(ws2, min_col=2, max_col=6, min_row=3)
    radar.set_categories(cats)

    colors = [BLUE, GREEN, ORANGE, RED, "8E44AD"]
    for i in range(min(5, len(df))):
        row_num = 4 + i
        values = Reference(ws2, min_col=2, max_col=6, min_row=row_num)
        radar.add_data(values, from_rows=True)
        radar.series[i].tx = SeriesLabel(v=ws2.cell(row=row_num, column=1).value)
        radar.series[i].graphicalProperties.solidFill = colors[i]

    ws2.add_chart(radar, f"A{comp_last + 2}")

    # ============================================================
    # SHEET 3: Eurostat Historical
    # ============================================================
    ws3 = wb.create_sheet("Eurostat Data")
    ws3.sheet_properties.tabColor = ORANGE

    ws3.merge_cells("A1:F1")
    ws3["A1"] = "Eurostat Tourism Data (NUTS-2)"
    ws3["A1"].font = title_font
    ws3["A1"].alignment = Alignment(horizontal="center")

    euro_headers = ["NUTS-2", "Year", "Arrivals", "Nights", "Avg Stay", "YoY Arrivals %"]
    for col_idx, h in enumerate(euro_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    if not df_euro.empty:
        for row_idx, (_, row) in enumerate(df_euro.iterrows(), 4):
            vals = [row["nuts2"], row["year"], row.get("arrivals"), row.get("nights"),
                    row.get("avg_stay"), row.get("yoy_arrivals_%")]
            for col_idx, v in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=col_idx)
                if pd.isna(v):
                    cell.value = ""
                elif isinstance(v, float) and col_idx in (3, 4):
                    cell.value = round(v)
                    cell.number_format = "#,##0"
                elif isinstance(v, float):
                    cell.value = round(v, 2)
                else:
                    cell.value = v
                cell.alignment = center
                cell.border = thin_border

    for i, w in enumerate([10, 8, 16, 18, 10, 16], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # SHEET 4: Shannon Entropy Weights
    # ============================================================
    ws4 = wb.create_sheet("Shannon Entropy")
    ws4.sheet_properties.tabColor = "8E44AD"

    weights = df.attrs.get("weights", WEIGHTS_FALLBACK)
    entropy_result = df.attrs.get("entropy_result")
    method = df.attrs.get("weighting_method", "Static")

    ws4.merge_cells("A1:F1")
    ws4["A1"] = f"Shannon Entropy Weighting ({method})"
    ws4["A1"].font = title_font
    ws4["A1"].alignment = Alignment(horizontal="center")

    # Entropy diagnostics table
    ent_headers = ["Component", "E_j (Entropy)", "d_j (Diversity)", "Weight w_j", "Data Source"]
    for col_idx, h in enumerate(ent_headers, 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sources_map = {
        "arrivals": "Eurostat tour_occ_arn2 (NUTS-2, annual)",
        "nights": "Eurostat tour_occ_nin2 (NUTS-2, annual)",
        "demand": "Google Trends API (weekly, GR)",
        "weather": "Open-Meteo Forecast API (hourly)",
        "avg_stay": "Derived: nights / arrivals",
    }
    for row_idx, var in enumerate(weights.keys(), 4):
        ws4.cell(row=row_idx, column=1, value=COMPONENT_NAMES.get(var, var)).border = thin_border

        if entropy_result:
            cell_e = ws4.cell(row=row_idx, column=2, value=entropy_result["entropy"][var])
            cell_e.number_format = "0.0000"
            cell_e.alignment = center
            cell_e.border = thin_border

            cell_d = ws4.cell(row=row_idx, column=3, value=entropy_result["diversity"][var])
            cell_d.number_format = "0.0000"
            cell_d.alignment = center
            cell_d.border = thin_border
        else:
            ws4.cell(row=row_idx, column=2, value="N/A").alignment = center
            ws4.cell(row=row_idx, column=3, value="N/A").alignment = center

        cell_w = ws4.cell(row=row_idx, column=4, value=float(weights[var]))
        cell_w.number_format = "0.00%"
        cell_w.alignment = center
        cell_w.border = thin_border
        # Highlight highest weight
        if var == max(weights, key=weights.get):
            cell_w.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ws4.cell(row=row_idx, column=5, value=sources_map.get(var, "")).border = thin_border

    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 16
    ws4.column_dimensions["D"].width = 14
    ws4.column_dimensions["E"].width = 45

    # Formulas section
    r = 4 + len(weights) + 2
    ws4.cell(row=r, column=1, value="Shannon Entropy Formula:").font = Font(bold=True, color=BLUE)
    formulas = [
        "p_ij = x_ij / Sum_i(x_ij)                     -- proportion of municipality i in variable j",
        "E_j = -(1/ln(n)) * Sum_i[p_ij * ln(p_ij)]     -- entropy (0=max info, 1=no info)",
        "d_j = 1 - E_j                                   -- diversification degree",
        "w_j = d_j / Sum_k(d_k)                          -- normalized weight (Sum w = 1)",
    ]
    for i, f in enumerate(formulas):
        ws4.cell(row=r + 1 + i, column=1, value=f).font = Font(name="Consolas", size=10)

    r2 = r + len(formulas) + 2
    ws4.cell(row=r2, column=1, value="Composite Score:").font = Font(bold=True, color=BLUE)
    ws4.cell(row=r2 + 1, column=1, value="Score = Sum_j(w_j * norm_j)").font = Font(name="Consolas", size=10)
    ws4.cell(row=r2 + 2, column=1, value="where norm_j is min-max normalized to [0, 100]").font = Font(italic=True, size=10)

    r3 = r2 + 4
    ws4.cell(row=r3, column=1, value="Interpretation:").font = Font(bold=True, color=BLUE)
    max_var = max(weights, key=weights.get)
    min_var = min(weights, key=weights.get)
    ws4.cell(row=r3 + 1, column=1,
             value=f"Highest weight: {COMPONENT_NAMES[max_var]} ({weights[max_var]:.1%}) -- most heterogeneous across municipalities")
    ws4.cell(row=r3 + 2, column=1,
             value=f"Lowest weight: {COMPONENT_NAMES[min_var]} ({weights[min_var]:.1%}) -- most uniform (least discriminating)")

    ws4.cell(row=r3 + 4, column=1, value="Generated:").font = Font(bold=True, color=BLUE)
    ws4.cell(row=r3 + 4, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    # Save
    wb.save(output_path)
    log.info(f"Excel saved: {output_path}")


def main():
    log.info("Fetching real data...")
    df_euro = fetch_eurostat_tourism()
    df_trends = fetch_google_trends()
    df_weather = fetch_weather()

    df = build_resilience_snapshot(df_euro, df_trends, df_weather)

    output = "ResilienceIQ_Dashboard.xlsx"
    build_excel(df, df_euro, output)
    print(f"\n[OK] Dashboard saved: {output}")
    print(f"     Municipalities: {len(df)}")
    print(f"     Eurostat year: {df['eurostat_year'].iloc[0]}")
    print(f"     Demand index: {df['demand_index'].iloc[0]}")
    print(f"     Score range: {df['resilience_score'].min():.1f} - {df['resilience_score'].max():.1f}")


if __name__ == "__main__":
    main()
